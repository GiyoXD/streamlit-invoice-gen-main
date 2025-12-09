"""
Script to reconstruct an Excel template from its JSON "DNA" (template config).

Usage:
    python scripts/reconstruct_template.py <path_to_template_config.json> <output_path.xlsx>
"""

import sys
import json
import argparse
import base64
from io import BytesIO
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, Color
from openpyxl.drawing.image import Image as XLImage

def reconstruct_template(config_path: str, output_path: str):
    config_path = Path(config_path)
    output_path = Path(output_path)
    
    if not config_path.exists():
        print(f"Error: Config file not found: {config_path}")
        return
        
    print(f"Loading DNA from: {config_path}")
    with open(config_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    if "template_layout" not in data:
        print("Error: Invalid config format. Missing 'template_layout' key.")
        return
        
    layout_data = data["template_layout"]
    
    # Create blank workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    print(f"Reconstructing {len(layout_data)} sheets...")
    
    for sheet_name, sheet_data in layout_data.items():
        print(f"  Building sheet: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        
        # Column Widths
        if "col_widths" in sheet_data:
             for letter, width in sheet_data["col_widths"].items():
                 ws.column_dimensions[letter].width = float(width)
        
        # --- 1. Header Reconstruction ---
        print("    Restoring Header...")
        
        # Row Heights
        if "header_row_heights" in sheet_data:
            for row_idx, height in sheet_data["header_row_heights"].items():
                ws.row_dimensions[int(row_idx)].height = float(height)
                
        # Content
        if "header_content" in sheet_data:
            for coord, value in sheet_data["header_content"].items():
                ws[coord] = value
                
        # Styles (Header)
        if "header_styles" in sheet_data:
             for coord, style_data in sheet_data["header_styles"].items():
                 apply_style_from_dict(ws[coord], style_data)
        
        # Images (Header)
        if "header_images" in sheet_data:
            restore_images(ws, sheet_data["header_images"])

        # Merges
        if "header_merges" in sheet_data:
            for merge_range in sheet_data["header_merges"]:
                try:
                    ws.merge_cells(merge_range)
                except Exception as e:
                    print(f"      Warning: Failed to merge {merge_range}: {e}")

        # --- 2. Footer Reconstruction ---
        print("    Restoring Footer...")
        
        # Row Heights
        if "footer_row_heights" in sheet_data:
            for row_idx, height in sheet_data["footer_row_heights"].items():
                ws.row_dimensions[int(row_idx)].height = float(height)
                
        # Content
        if "footer_content" in sheet_data:
            for coord, value in sheet_data["footer_content"].items():
                ws[coord] = value
                
        # Styles (Footer)
        if "footer_styles" in sheet_data:
             for coord, style_data in sheet_data["footer_styles"].items():
                 apply_style_from_dict(ws[coord], style_data)
                 
        # Images (Footer)
        if "footer_images" in sheet_data:
            restore_images(ws, sheet_data["footer_images"])
                
        # Merges
        if "footer_merges" in sheet_data:
            for merge_range in sheet_data["footer_merges"]:
                try:
                    ws.merge_cells(merge_range)
                except Exception as e:
                    print(f"      Warning: Failed to merge {merge_range}: {e}")
                    
    # Save
    print(f"Saving reconstructed template to: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print("Done!")

def restore_images(ws, images_list):
    """Restore images from base64 list."""
    if not images_list: return
    
    for img_data in images_list:
        try:
            if not img_data.get("data"): continue
            
            # Decode
            data = base64.b64decode(img_data["data"])
            img = XLImage(BytesIO(data))
            
            # Set Anchor
            if "anchor" in img_data:
                img.anchor = img_data["anchor"]
                
            ws.add_image(img)
            
        except Exception as e:
            print(f"      Warning: Failed to restore image: {e}")

def apply_style_from_dict(cell, style_dict):
    """Apply style dict to a cell."""
    if not style_dict: return
    
    # Font
    if "font" in style_dict:
        f = style_dict["font"]
        cell.font = Font(
            name=f.get("name"),
            size=f.get("size"),
            bold=f.get("bold"),
            italic=f.get("italic"),
            color=f.get("color") # Warning: Hex string usually works if not None
        )
        
    # Alignment
    if "alignment" in style_dict:
        a = style_dict["alignment"]
        cell.alignment = Alignment(
            horizontal=a.get("horizontal"),
            vertical=a.get("vertical"),
            wrap_text=a.get("wrap_text")
        )
        
    # Border
    if "border" in style_dict and style_dict["border"]:
        b = style_dict["border"]
        cell.border = Border(
            left=Side(style=b.get("left")) if b.get("left") else None,
            right=Side(style=b.get("right")) if b.get("right") else None,
            top=Side(style=b.get("top")) if b.get("top") else None,
            bottom=Side(style=b.get("bottom")) if b.get("bottom") else None
        )
        
    # Fill
    if "fill" in style_dict:
        fl = style_dict["fill"]
        if fl.get("type") == "solid" and fl.get("color"):
             cell.fill = PatternFill(start_color=fl.get("color"), end_color=fl.get("color"), fill_type="solid")
             
    # Number Format
    if "number_format" in style_dict:
        cell.number_format = style_dict["number_format"]

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
        
    reconstruct_template(sys.argv[1], sys.argv[2])
