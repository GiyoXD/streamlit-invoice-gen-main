import os
import json
import sys
from pathlib import Path
import openpyxl

# Add core to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from core.invoice_generator.generate_invoice import run_invoice_generation

def create_test_assets(base_dir):
    base_dir = Path(base_dir)
    base_dir.mkdir(exist_ok=True, parents=True)

    # 1. Create Data
    data = {
        "INVOICE_NUM": "INV-TEST-001",
        "CUSTOMER": "Test Customer",
        "AMOUNT": "999.99"
    }
    data_path = base_dir / "PlaceholderTest.json"
    with open(data_path, 'w') as f:
        json.dump(data, f)

    # 2. Create Config
    config = {
        "processing": {
            "sheets": ["Sheet1"],
            "data_sources": {
                "Sheet1": "placeholder"
            }
        },
        "data_bundle": {
            "Sheet1": {
                "mapping": {} 
            }
        }
    }
    config_path = base_dir / "PlaceholderTest_bundle_config.json"
    with open(config_path, 'w') as f:
        json.dump(config, f)

    # 3. Create Template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = "Invoice Number:"
    ws['B1'] = "{{INVOICE_NUM}}"
    ws['A2'] = "Customer:"
    ws['B2'] = "{{CUSTOMER}}"
    ws['A3'] = "Total:"
    ws['B3'] = "{{AMOUNT}}"
    
    template_path = base_dir / "PlaceholderTest.xlsx"
    wb.save(template_path)
    
    return data_path, config_path, template_path

def verify():
    test_dir = Path("tests/temp_placeholder_test")
    data_path, config_path, template_path = create_test_assets(test_dir)
    
    output_path = test_dir / "PlaceholderResult.xlsx"
    
    print(f"Running generation...")
    print(f"Data: {data_path}")
    print(f"Config: {config_path}")
    print(f"Template: {template_path}")
    
    try:
        run_invoice_generation(
            input_data_path=data_path,
            output_path=output_path,
            template_dir=test_dir, # Use test dir as template dir
            config_dir=test_dir,   # Use test dir as config dir
            explicit_config_path=config_path,
            explicit_template_path=template_path
        )
        
        print("Generation completed. Verifying output...")
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Sheet1"]
        
        val_inv = ws['B1'].value
        val_cust = ws['B2'].value
        val_amt = ws['B3'].value
        
        print(f"B1: {val_inv}")
        print(f"B2: {val_cust}")
        print(f"B3: {val_amt}")
        
        if val_inv == "INV-TEST-001" and val_cust == "Test Customer" and val_amt == "999.99":
            print("SUCCESS: All placeholders replaced correctly.")
        else:
            print("FAILURE: Values do not match expected output.")
            
    except Exception as e:
        print(f"FAILURE: Exception occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify()
