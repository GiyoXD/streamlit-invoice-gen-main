import os
import json
import sys
from pathlib import Path
import openpyxl

# Add core to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from core.invoice_generator.generate_invoice import run_invoice_generation

def create_test_assets(base_dir):
    base_dir = Path(base_dir)
    base_dir.mkdir(exist_ok=True, parents=True)

    # 1. Create Data
    data = {
        "INVOICE_NUM": "ALIAS-TEST-001",
        "CUSTOMER": "Alias Customer",
        "AMOUNT": "555.55"
    }
    data_path = base_dir / "AliasTest.json"
    with open(data_path, 'w') as f:
        json.dump(data, f)

    # 2. Create Config using NEW ALIASES
    config = {
        "processing": {
            "processing_order": ["Sheet1"],  # ALIAS for 'sheets'
            "sheet_processing_types": {      # ALIAS for 'data_sources'
                "Sheet1": "placeholder"
            }
        },
        "data_bundle": {
            "Sheet1": {
                "mapping": {} 
            }
        }
    }
    config_path = base_dir / "AliasTest_bundle_config.json"
    with open(config_path, 'w') as f:
        json.dump(config, f)

    # 3. Create Template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = "Invoice Number:"
    ws['B1'] = "{{INVOICE_NUM}}"
    
    template_path = base_dir / "AliasTest.xlsx"
    wb.save(template_path)
    
    return data_path, config_path, template_path

def verify():
    test_dir = Path("tests/temp_alias_test")
    data_path, config_path, template_path = create_test_assets(test_dir)
    
    output_path = test_dir / "AliasResult.xlsx"
    
    print(f"Running generation with ALIASES...")
    
    try:
        run_invoice_generation(
            input_data_path=data_path,
            output_path=output_path,
            template_dir=test_dir,
            config_dir=test_dir,
            explicit_config_path=config_path,
            explicit_template_path=template_path
        )
        
        print("Generation completed. Verifying output...")
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Sheet1"]
        val_inv = ws['B1'].value
        
        print(f"B1: {val_inv}")
        
        if val_inv == "ALIAS-TEST-001":
            print("SUCCESS: Aliases worked correctly.")
        else:
            print("FAILURE: Values do not match expected output.")
            
    except Exception as e:
        print(f"FAILURE: Exception occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify()
