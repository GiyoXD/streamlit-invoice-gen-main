
import openpyxl
from openpyxl.styles import Font
import os

def create_valid_template(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Metadata (Rows 1-4)
    ws["A1"] = "INVOICE NO:"
    ws["B1"] = "JFINV-001"
    ws["A2"] = "DATE:"
    ws["B2"] = "2023-10-01"
    
    # Header (Row 5)
    header_row = 5
    headers = ["PO NO", "ITEM NO", "DESC", "QTY", "PRICE", "AMOUNT"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        
    # Data (Rows 6-8)
    data = [
        ["PO123", "ITM01", "Leather", 100, 5.0, 500.0],
        ["PO123", "ITM02", "Leather", 200, 5.0, 1000.0],
        ["PO124", "ITM03", "Leather", 150, 5.0, 750.0],
    ]
    
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
            
    # Footer (Row 10 - leaves one empty row)
    footer_row = header_row + len(data) + 2
    ws.cell(row=footer_row, column=1, value="TOTAL:")
    ws.cell(row=footer_row, column=6, value="=SUM(F6:F8)")
    
    # Add a styled footer below to ensure footer detection logic works locally
    ws.cell(row=footer_row+1, column=1, value="Signature")
    
    # Output directory
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"Created valid template at {filename}")

if __name__ == "__main__":
    create_valid_template("tests/database/hostfile/Valid_Template.xlsx")
