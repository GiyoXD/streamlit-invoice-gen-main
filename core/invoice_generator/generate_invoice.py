# core/invoice_generator/generate_invoice.py
import os
import json
import pickle
import argparse
import sys
import time
import datetime
import logging
import traceback
from pathlib import Path
from typing import Optional, Dict, Any, List
import openpyxl
import re
import ast

# Keep your existing imports
from .config.config_loader import BundledConfigLoader
from .builders.workbook_builder import WorkbookBuilder
from .processors.single_table_processor import SingleTableProcessor
from .processors.multi_table_processor import MultiTableProcessor

logger = logging.getLogger(__name__)

# --- Helper Functions ---

def derive_paths(input_data_path: str, template_dir: str, config_dir: str) -> Optional[Dict[str, Path]]:
    """
    Derive paths for config and template based on input data filename.
    """
    input_path = Path(input_data_path)
    stem = input_path.stem
    
    # Prioritize bundle config to avoid picking up data file as config
    config_path = Path(config_dir) / f"{stem}_bundle_config.json"
    if not config_path.exists():
        config_path = Path(config_dir) / f"{stem}.json"
    
    # Fallback to default config if specific not found
    if not config_path.exists():
        default_config = Path(config_dir) / "default.json"
        if default_config.exists():
             config_path = default_config
        else:
             # If no config found, we can't proceed unless we have a strategy
             pass

    # Template path - ideally derived from config, but we need config first.
    # Strategy: Load config, check for template name. If not, use stem.
    template_path = None
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
                # Check if template is specified in meta or processing
                template_name = cfg.get('_meta', {}).get('template_name')
                if template_name:
                    template_path = Path(template_dir) / template_name
        except:
            pass
    
    if not template_path:
        template_path = Path(template_dir) / f"{stem}.xlsx"
        if not template_path.exists():
             # Fallback to generic Invoice.xlsx
             fallback = Path(template_dir) / "Invoice.xlsx"
             if fallback.exists():
                 template_path = fallback

    if config_path.exists() and template_path and template_path.exists():
        return {
            'data': input_path,
            'config': config_path,
            'template': template_path
        }
    
    logger.error(f"Could not derive paths. Config: {config_path} (Exists: {config_path.exists()}), Template: {template_path} (Exists: {template_path and template_path.exists()})")
    return None

def load_data(path: Path) -> Dict[str, Any]:
    """Load invoice data from JSON."""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Failed to load data from {path}: {e}")
        return {}

def generate_metadata(output_path: Path, status: str, duration: float, 
                     processed: List[str], failed: List[str], error_msg: Optional[str],
                     invoice_data: Dict[str, Any], args: argparse.Namespace,
                     replacements: List[Any], header_info: Dict[str, Any]) -> None:
    """Generate metadata JSON file side-by-side with output."""
    meta_path = output_path.parent / f"{output_path.stem}_metadata.json"
    
    metadata = {
        "status": status,
        "timestamp": datetime.datetime.now().isoformat(),
        "duration_seconds": duration,
        "output_file": str(output_path.name),
        "sheets_processed": processed,
        "sheets_failed": failed,
        "error_message": error_msg,
        "input_metadata": invoice_data.get("metadata", {}),
        "generation_args": vars(args) if args else {},
        # "replacements_log": replacements, # Can be verbose
        # "header_info": header_info
    }
    
    try:
        with open(meta_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=4)
    except Exception as e:
        logger.error(f"Failed to write metadata: {e}")

def run_invoice_generation(
    input_data_path: Path,
    output_path: Path,
    template_dir: Path,
    config_dir: Path,
    daf_mode: bool = False,
    custom_mode: bool = False
) -> Path:
    """
    Library entry point for invoice generation. 
    Raises exceptions instead of sys.exit() for better control in the Orchestrator.
    """
    start_time = time.time()
    
    # Ensure paths are Path objects
    input_data_path = Path(input_data_path).resolve()
    output_path = Path(output_path).resolve()
    template_dir = Path(template_dir).resolve()
    config_dir = Path(config_dir).resolve()

    logger.info("=== Starting Invoice Generation (Library Call) ===")
    logger.debug(f"Input: {input_data_path}, Output: {output_path}")

    # 1. Derive Paths
    paths = derive_paths(str(input_data_path), str(template_dir), str(config_dir))
    if not paths:
        raise FileNotFoundError(f"Could not derive template/config paths for {input_data_path.name}")

    # 2. Load Configuration
    try:
        config_loader = BundledConfigLoader(paths['config'])
    except Exception as e:
        raise RuntimeError(f"Failed to load configuration: {e}") from e
    
    # 3. Load Data
    invoice_data = load_data(paths['data'])
    if not invoice_data:
        raise RuntimeError(f"Failed to load input data from {paths['data']}")

    # 4. Prepare Output Directory
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    template_workbook = None
    output_workbook = None
    processing_successful = True
    sheets_processed = []
    sheets_failed = []
    
    # Capture these for metadata later
    replacements_log = []
    header_info = {}

    try:
        # Step 5: Load Template & Create Workbook
        logger.info(f"Loading template from: {paths['template']}")
        template_workbook = openpyxl.load_workbook(paths['template'], read_only=False)
        
        workbook_builder = WorkbookBuilder(sheet_names=template_workbook.sheetnames)
        output_workbook = workbook_builder.build()
        
        # Step 6: Determine Sheets to Process
        sheets_to_process_config = config_loader.get_sheets_to_process()
        sheets_to_process = [s for s in sheets_to_process_config if s in output_workbook.sheetnames]

        if not sheets_to_process:
            raise ValueError("No valid sheets found to process in configuration.")

        # Global calculation (legacy support)
        final_grand_total_pallets = 0
        processed_tables = invoice_data.get('processed_tables_data', {})
        if isinstance(processed_tables, dict):
            final_grand_total_pallets = sum(
                int(c) for t in processed_tables.values() 
                for c in t.get("pallet_count", []) 
                if str(c).isdigit()
            )

        # Step 7: Processing Loop
        for sheet_name in sheets_to_process:
            logger.info(f"Processing sheet '{sheet_name}'")
            
            template_worksheet = template_workbook[sheet_name]
            output_worksheet = output_workbook[sheet_name]
            
            sheet_config = config_loader.get_sheet_config(sheet_name)
            data_source_indicator = config_loader.get_data_source_type(sheet_name)

            if not data_source_indicator:
                logger.warning(f"Skipping '{sheet_name}': No data source configured.")
                continue

            # Instantiate Processor
            # Mock CLI args for compatibility with existing processors if they rely on it
            # Ideally, refactor processors to take booleans, but this shim works for now.
            mock_args = argparse.Namespace(DAF=daf_mode, custom=custom_mode)

            processor = None
            if data_source_indicator in ["processed_tables_multi", "processed_tables"]:
                processor = MultiTableProcessor(
                    template_workbook=template_workbook,
                    output_workbook=output_workbook,
                    template_worksheet=template_worksheet,
                    output_worksheet=output_worksheet,
                    sheet_name=sheet_name,
                    sheet_config=sheet_config,
                    config_loader=config_loader,
                    data_source_indicator=data_source_indicator,
                    invoice_data=invoice_data,
                    cli_args=mock_args, 
                    final_grand_total_pallets=final_grand_total_pallets
                )
            else:
                processor = SingleTableProcessor(
                    template_workbook=template_workbook,
                    output_workbook=output_workbook,
                    template_worksheet=template_worksheet,
                    output_worksheet=output_worksheet,
                    sheet_name=sheet_name,
                    sheet_config=sheet_config,
                    config_loader=config_loader,
                    data_source_indicator=data_source_indicator,
                    invoice_data=invoice_data,
                    cli_args=mock_args,
                    final_grand_total_pallets=final_grand_total_pallets
                )

            if processor:
                if processor.process():
                    sheets_processed.append(sheet_name)
                    # Collect logs
                    if hasattr(processor, 'replacements_log'):
                        replacements_log.extend(processor.replacements_log)
                    if hasattr(processor, 'header_info'):
                        header_info.update(processor.header_info)
                else:
                    sheets_failed.append(sheet_name)
                    processing_successful = False

        # Step 8: Save
        logger.info(f"Saving workbook to {output_path}")
        output_workbook.save(output_path)

    except Exception as e:
        logger.error(f"Error during generation: {e}")
        traceback.print_exc()
        processing_successful = False
        if output_workbook:
            # Attempt emergency save
            try: 
                output_workbook.save(output_path)
            except: 
                pass
        raise e # Re-raise to let Orchestrator know
    finally:
        if template_workbook: template_workbook.close()
        if output_workbook: output_workbook.close()

    total_time = time.time() - start_time
    
    # Generate Metadata
    status = "success" if processing_successful and not sheets_failed else "error"
    error_msg = f"Failed sheets: {sheets_failed}" if sheets_failed else None
    
    # Mock CLI args for metadata generation
    meta_args = argparse.Namespace(
        DAF=daf_mode, 
        custom=custom_mode, 
        input_data_file=str(input_data_path), 
        configdir=str(config_dir)
    )
    
    generate_metadata(
        output_path, status, total_time, sheets_processed, sheets_failed, 
        error_msg, invoice_data, meta_args, replacements_log, header_info
    )

    return output_path

def main():
    """CLI Entry point for backward compatibility."""
    parser = argparse.ArgumentParser(description="Generate Invoice CLI")
    parser.add_argument("input_data_file", help="Path to input data file")
    parser.add_argument("-o", "--output", default="result.xlsx", help="Output path")
    parser.add_argument("-t", "--templatedir", default="./TEMPLATE", help="Template dir")
    parser.add_argument("-c", "--configdir", default="./configs", help="Config dir")
    parser.add_argument("--DAF", action="store_true", help="DAF mode")
    parser.add_argument("--custom", action="store_true", help="Custom mode")
    parser.add_argument("--debug", action="store_true", help="Debug logging")
    
    args = parser.parse_args()
    
    # Configure Logging for CLI
    level = logging.DEBUG if args.debug else logging.INFO
    logging.basicConfig(level=level, format='%(levelname)s: %(message)s')
    
    try:
        run_invoice_generation(
            Path(args.input_data_file),
            Path(args.output),
            Path(args.templatedir),
            Path(args.configdir),
            daf_mode=args.DAF,
            custom_mode=args.custom
        )
        print(f"Successfully generated: {args.output}")
    except Exception as e:
        print(f"Generation failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()