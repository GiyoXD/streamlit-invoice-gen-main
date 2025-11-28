import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter

from ..utils.merge_utils import store_original_merges

logger = logging.getLogger(__name__)

class TemplatePreprocessor:
    """
    Pre-processes the Excel template to:
    1. Replace specific text patterns (dates, invoice nos) with placeholders (JFTIME, JFINV).
    2. Remove dummy data rows to create a clean template for the generator.
    """

    def __init__(self):
        self.replacement_patterns = {
            'date': {
                'patterns': [
                    r'date[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',
                    r'date[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}',
                    r'date[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{2}',
                    r'dated[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',
                    r'dated[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}',
                    r'dated[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{2}'
                ],
                'replacement': 'JFTIME'
            },
            'date_label': {
                'patterns': [
                    r'date[^a-zA-Z]',
                    r'dated[^a-zA-Z]',
                    r'invoice\s+date[^a-zA-Z]',
                    r'contract\s+date[^a-zA-Z]'
                ],
                'replacement': 'JFTIME'
            },
            'invoice_label': {
                'patterns': [
                    r'invoice[^a-zA-Z]*no[^a-zA-Z]',
                    r'inv[^a-zA-Z]*no[^a-zA-Z]',
                    r'bill[^a-zA-Z]*no[^a-zA-Z]'
                ],
                'replacement': 'JFINV'
            },
            'ref_label': {
                'patterns': [
                    r'ref[^a-zA-Z]*no[^a-zA-Z]',
                    r'reference[^a-zA-Z]*no[^a-zA-Z]',
                    r'our[^a-zA-Z]*ref[^a-zA-Z]'
                ],
                'replacement': 'JFREF'
            },
            'contract_no': {
                'patterns': [
                    r'contract[^a-zA-Z]*no[^a-zA-Z]*[\w\-/]+',
                    r'contract[^a-zA-Z]*[\w\-/]+',
                    r'cont[^a-zA-Z]*no[^a-zA-Z]*[\w\-/]+'
                ],
                'replacement': 'JFINV'
            },
            'invoice_no': {
                'patterns': [
                    r'invoice[^a-zA-Z]*no[^a-zA-Z]*\d+\-?\d*',
                    r'inv[^a-zA-Z]*no[^a-zA-Z]*\d+\-?\d*',
                    r'bill[^a-zA-Z]*no[^a-zA-Z]*\d+\-?\d*'
                ],
                'replacement': 'JFINV'
            },
            'ref_no': {
                'patterns': [
                    r'ref[^a-zA-Z]*no[^a-zA-Z]*[A-Z]+\d+-\d+',
                    r'reference[^a-zA-Z]*no[^a-zA-Z]*[A-Z]+\d+-\d+',
                    r'our[^a-zA-Z]*ref[^a-zA-Z]*[A-Z]+\d+-\d+'
                ],
                'replacement': 'JFREF'
            },
            'etd': {
                'patterns': [
                    r'etd[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',
                    r'etd[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}',
                    r'etd[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{2}',
                    r'estimated[^a-zA-Z]*time[^a-zA-Z]*of[^a-zA-Z]*departure[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',
                    r'departure[^a-zA-Z]*\d{1,2}/\d{1,2}/\d{4}',
                    r'departure[^a-zA-Z]*\d{4}/\d{1,2}/\d{1,2}'
                ],
                'replacement': 'JFTIME'
            }
        }
        
        self.header_keywords = [
            "P.O", "ITEM", "Description", "Quantity", "Amount", "Unit Price", 
            "Mark", "No", "Pallet", "N.W", "G.W", "CBM", "PCS", "SF"
        ]
        
        self.formula_patterns = [
            r'=sum\(',
            r'=SUM\(',
            r'=Sum\('
        ]

    def process(self, template_path: Path) -> bool:
        """
        Process the template file in-place.
        """
        try:
            logger.info(f"Pre-processing template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # 1. Text Replacement (Safe operation)
            logger.info("Performing text replacements...")
            self._apply_text_replacements(workbook)
            
            # 2. Clear Data Rows (Non-destructive to structure)
            # Instead of deleting rows (which breaks merges), we just clear the content.
            logger.info("Clearing dummy data rows...")
            self._clear_data_rows(workbook)
            
            # 3. Unhide rows (Safe operation)
            self._unhide_all_rows(workbook)
            
            workbook.save(template_path)
            logger.info(f"Template pre-processing completed: {template_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error pre-processing template: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _apply_text_replacements(self, workbook: openpyxl.Workbook):
        """Apply text replacements using circular pattern checking."""
        processed_cells = set()
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            logger.info(f"Processing text replacements in sheet: {sheet_name}")
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        cell_coord = cell.coordinate
                        if cell_coord in processed_cells:
                            continue
                            
                        label_match = self._find_label_match(cell.value)
                        if label_match:
                            if self._is_likely_table_data_area(worksheet, cell):
                                continue
                                
                            target_cell = self._find_target_cell_circular(worksheet, cell, label_match)
                            if target_cell and target_cell.coordinate not in processed_cells:
                                logger.info(f"Replacing '{target_cell.value}' with '{label_match['replacement']}' at {target_cell.coordinate}")
                                target_cell.value = label_match['replacement']
                                processed_cells.add(cell_coord)
                                processed_cells.add(target_cell.coordinate)

    def _find_label_match(self, text: str) -> Optional[Dict[str, str]]:
        for category, config in self.replacement_patterns.items():
            for pattern in config['patterns']:
                if re.search(pattern, text, re.IGNORECASE):
                    return {'category': category, 'replacement': config['replacement']}
        return None

    def _find_target_cell_circular(self, worksheet: Worksheet, label_cell: Cell, label_match: Dict[str, str]) -> Optional[Cell]:
        row = label_cell.row
        col = label_cell.column
        circular_pattern = [
            (0, 1, 10), (0, 2, 8), (1, 0, 7), (-1, 0, 6),
            (1, 1, 5), (-1, 1, 5), (0, -1, 4), (1, -1, 3),
            (-1, -1, 3), (0, 3, 2), (2, 0, 2), (-2, 0, 2)
        ]
        
        candidates = []
        for r_off, c_off, priority in circular_pattern:
            t_row, t_col = row + r_off, col + c_off
            if t_row < 1 or t_col < 1: continue
            try:
                cell = worksheet.cell(row=t_row, column=t_col)
                score = self._evaluate_target_cell(cell, label_match, priority)
                if score > 0:
                    candidates.append({'cell': cell, 'score': score})
            except: continue
            
        if candidates:
            return max(candidates, key=lambda x: x['score'])['cell']
        return None

    def _evaluate_target_cell(self, cell: Cell, label_match: Dict[str, str], priority: int) -> int:
        if cell.value is None: return 0
        val = str(cell.value).strip()
        if not val: return 0
        
        score = max(1, priority // 2)
        cat = label_match['category']
        
        # Simplified scoring logic
        if cat in ['date', 'etd', 'date_label']:
            if re.search(r'\d{4}.*\d{1,2}.*\d{1,2}', val) or re.search(r'\d{1,2}.*\d{1,2}.*\d{2,4}', val):
                score += 30
            else: return 0
        elif cat in ['invoice_no', 'invoice_label', 'contract_no']:
            if re.search(r'[A-Z].*\d', val) or re.search(r'\d{4,}', val):
                score += 25
            else: return 0
        elif cat in ['ref_no', 'ref_label']:
            if re.search(r'[A-Z].*\d', val):
                score += 25
            else: return 0
            
        return score

    def _is_likely_table_data_area(self, worksheet: Worksheet, cell: Cell) -> bool:
        # Simplified check
        row, col = cell.row, cell.column
        indicators = 0
        for r in range(row-1, row+2):
            for c in range(col-2, col+3):
                try:
                    val = str(worksheet.cell(row=r, column=c).value)
                    if re.search(r'^\d+\.?\d*$', val) or re.search(r'qty|price|amount|total', val, re.I):
                        indicators += 1
                except: pass
        return indicators >= 4

    def _clear_data_rows(self, workbook: openpyxl.Workbook):
        """
        Clears content from data rows without deleting them.
        This preserves merges and structure.
        """
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            logger.info(f"Clearing data rows in sheet: {sheet_name}")
            
            header_rows = self._find_header_rows(worksheet)
            for header_row in header_rows:
                self._clear_table_data(worksheet, header_row)

    def _find_header_rows(self, worksheet: Worksheet) -> List[int]:
        header_rows = []
        for row in range(1, min(50, worksheet.max_row + 1)):
            matches = 0
            total = 0
            for col in range(1, worksheet.max_column + 1):
                val = worksheet.cell(row=row, column=col).value
                if val:
                    total += 1
                    if any(k.lower() in str(val).lower() for k in self.header_keywords):
                        matches += 1
            if total > 0 and (matches / total) >= 0.4:
                header_rows.append(row)
        return header_rows

    def _clear_table_data(self, worksheet: Worksheet, header_row: int):
        """Identify table bounds and clear content between header and footer."""
        formula_row = None
        # Find row with SUM formulas
        for row in range(header_row + 1, min(header_row + 51, worksheet.max_row + 1)):
            sums = 0
            for col in range(1, worksheet.max_column + 1):
                val = str(worksheet.cell(row=row, column=col).value)
                if any(re.search(p, val, re.I) for p in self.formula_patterns):
                    sums += 1
            if sums >= 1:
                formula_row = row
                break
        
        if formula_row:
            # Clear content from header_row + 1 to formula_row - 1
            start_row = header_row + 1
            end_row = formula_row - 1
            
            if start_row <= end_row:
                logger.info(f"Clearing content from rows {start_row} to {end_row}")
                for row in range(start_row, end_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if not isinstance(cell, MergedCell):
                            cell.value = None


    def _unhide_all_rows(self, workbook: openpyxl.Workbook):
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.row_dimensions.values():
                row.hidden = False
