"""
Template Analyzer - Extracts structure from Excel templates for auto config generation.

This module analyzes Excel template files to extract:
- Sheet names and structure
- Header row positions
- Column layouts (merged cells, widths)
- Font/style information
- Data source hints (aggregation vs processed_tables)
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class ColumnInfo:
    """Information about a single column."""
    id: str
    header: str
    col_index: int  # 1-based
    width: float
    format: str = "@"
    alignment: str = "center"
    rowspan: int = 1
    colspan: int = 1
    children: List['ColumnInfo'] = field(default_factory=list)
    wrap_text: bool = False


@dataclass
class SheetAnalysis:
    """Complete analysis of a single sheet."""
    name: str
    header_row: int
    columns: List[ColumnInfo]
    data_source: str  # "aggregation" or "processed_tables_multi"
    header_font: Dict[str, Any]
    data_font: Dict[str, Any]
    row_heights: Dict[str, float]  # "header", "data", "footer" -> height
    has_multi_row_header: bool = False
    static_content_hints: Dict[str, List[str]] = field(default_factory=dict)


@dataclass
class TemplateAnalysisResult:
    """Complete template analysis result."""
    file_path: str
    customer_code: str
    sheets: List[SheetAnalysis]
    

class TemplateAnalyzer:
    """Analyzes Excel templates to extract structure for config generation."""
    
    # Known header keywords for column ID mapping
    HEADER_MAPPINGS = {
        # Common headers -> column IDs
        "mark": "col_static",
        "mark & n": "col_static",
        "mark & no": "col_static",
        "p.o": "col_po",
        "po": "col_po",
        "p.o.": "col_po",
        "item": "col_item",
        "item n": "col_item",
        "description": "col_desc",
        "desc": "col_desc",
        "quantity": "col_qty_header",
        "qty": "col_qty_header",
        "pcs": "col_qty_pcs",
        "sf": "col_qty_sf",
        "sqft": "col_qty_sf",
        "unit price": "col_unit_price",
        "unit": "col_unit_price",
        "price": "col_unit_price",
        "amount": "col_amount",
        "total": "col_amount",
        "n.w": "col_net",
        "net": "col_net",
        "nw": "col_net",
        "g.w": "col_gross",
        "gross": "col_gross",
        "gw": "col_gross",
        "cbm": "col_cbm",
        "m3": "col_cbm",
        "no": "col_no",
        "no.": "col_no",
        "pallet": "col_pallet",
    }
    
    # Sheets that use aggregation data
    AGGREGATION_SHEETS = {"invoice", "contract"}
    
    # Sheets that use processed tables data
    PROCESSED_SHEETS = {"packing list", "packing", "pl"}
    
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def analyze_template(self, template_path: str) -> TemplateAnalysisResult:
        """
        Analyze an Excel template file and extract structure.
        
        Args:
            template_path: Path to the Excel template file
            
        Returns:
            TemplateAnalysisResult with complete analysis
        """
        path = Path(template_path)
        if not path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        # Extract customer code from filename (e.g., "CLW.xlsx" -> "CLW")
        customer_code = path.stem.upper()
        
        self.logger.info(f"Analyzing template: {path.name} (customer: {customer_code})")
        
        workbook = openpyxl.load_workbook(template_path, data_only=False)
        
        sheets = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            analysis = self._analyze_sheet(worksheet, sheet_name)
            if analysis:
                sheets.append(analysis)
        
        return TemplateAnalysisResult(
            file_path=str(path.absolute()),
            customer_code=customer_code,
            sheets=sheets
        )
    
    def _analyze_sheet(self, worksheet: Worksheet, sheet_name: str) -> Optional[SheetAnalysis]:
        """Analyze a single worksheet."""
        try:
            self.logger.info(f"  Analyzing sheet: {sheet_name}")
            
            # Find header row
            header_row, header_cells = self._find_header_row(worksheet)
            if not header_row:
                self.logger.warning(f"    No header row found in {sheet_name}")
                return None
            
            self.logger.info(f"    Header row: {header_row}")
            
            # Analyze columns
            columns = self._analyze_columns(worksheet, header_row, header_cells)
            self.logger.info(f"    Found {len(columns)} columns")
            # DETAILED DEBUG LOGGING
            self.logger.info(f"    --- Column Analysis for {sheet_name} ---")
            for col in columns:
                self.logger.info(f"      [Col {col.col_index}] ID={col.id} Header='{col.header}' Width={col.width:.1f}")
            self.logger.info(f"    ----------------------------------------")
            
            # Determine data source type
            data_source = self._determine_data_source(sheet_name, columns)
            self.logger.info(f"    Data source: {data_source}")
            
            # Extract font info
            header_font = self._extract_font_info(worksheet, header_row, 1)
            data_font = self._extract_font_info(worksheet, header_row + 1, 1)
            
            # Extract row heights
            row_heights = self._extract_row_heights(worksheet, header_row)
            
            # Check for multi-row headers
            has_multi_row = self._check_multi_row_header(worksheet, header_row)
            
            # Detect static content hints (like "Mark & Nº" column content)
            static_hints = self._detect_static_content(worksheet, header_row, columns)
            
            return SheetAnalysis(
                name=sheet_name,
                header_row=header_row,
                columns=columns,
                data_source=data_source,
                header_font=header_font,
                data_font=data_font,
                row_heights=row_heights,
                has_multi_row_header=has_multi_row,
                static_content_hints=static_hints
            )
            
        except Exception as e:
            self.logger.error(f"    Error analyzing {sheet_name}: {e}")
            return None
    
    def _find_header_row(self, worksheet: Worksheet) -> Tuple[Optional[int], List[Tuple[int, str]]]:
        """
        Find the header row by looking for known header keywords.
        
        Returns:
            Tuple of (row_number, list of (col_index, header_text))
        """
        best_row = None
        best_matches = []
        best_score = 0
        
        # Search first 30 rows
        for row in range(1, min(31, worksheet.max_row + 1)):
            matches = []
            for col in range(1, min(15, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row, column=col)
                value = self._get_cell_value(cell)
                if value:
                    # Check if this looks like a header
                    value_lower = value.lower().strip()
                    for keyword in self.HEADER_MAPPINGS:
                        if keyword in value_lower:
                            matches.append((col, value))
                            break
            
            # Score based on number of matches
            if len(matches) > best_score:
                best_score = len(matches)
                best_row = row
                best_matches = matches
            
            # Debug log for potential header rows
            if len(matches) > 0:
                self.logger.debug(f"    Row {row}: {len(matches)} matches -> {[m[1] for m in matches]}")
        
        # Need at least 3 header matches to be confident
        if best_score >= 3:
            return best_row, best_matches
        
        return None, []
    
    def _get_cell_value(self, cell) -> Optional[str]:
        """Get cell value, handling merged cells."""
        if isinstance(cell, MergedCell):
            return None
        if cell.value is None:
            return None
        return str(cell.value).strip()
    
    def _analyze_columns(self, worksheet: Worksheet, header_row: int, 
                         header_cells: List[Tuple[int, str]]) -> List[ColumnInfo]:
        """Analyze columns from header row."""
        columns = []
        processed_cols = set()
        
        # Get merged cell ranges for this row
        merged_ranges = []
        for merged in worksheet.merged_cells.ranges:
            if merged.min_row <= header_row <= merged.max_row:
                merged_ranges.append(merged)
        
        for col in range(1, worksheet.max_column + 1):
            if col in processed_cols:
                continue
            
            cell = worksheet.cell(row=header_row, column=col)
            value = self._get_cell_value(cell)
            
            if not value:
                # Check if this is part of a merged cell
                for merged in merged_ranges:
                    if merged.min_col <= col <= merged.max_col:
                        # Get value from top-left of merged range
                        value = self._get_cell_value(
                            worksheet.cell(row=merged.min_row, column=merged.min_col)
                        )
                        break
            
            if not value:
                continue
            
            # Determine column ID
            col_id = self._determine_column_id(value, col)
            
            # Get column width
            col_letter = get_column_letter(col)
            width = worksheet.column_dimensions[col_letter].width or 10
            
            # Check for merged cells (colspan/rowspan)
            colspan = 1
            rowspan = 1
            for merged in merged_ranges:
                if merged.min_col == col and merged.min_row == header_row:
                    colspan = merged.max_col - merged.min_col + 1
                    rowspan = merged.max_row - merged.min_row + 1
                    # Mark these columns as processed
                    for c in range(merged.min_col, merged.max_col + 1):
                        processed_cols.add(c)
                    break
            
            # Determine format based on column type
            format_str = self._determine_format(col_id, value)
            
            # Check alignment
            alignment = "center"
            if cell.alignment:
                alignment = cell.alignment.horizontal or "center"
            
            # Check wrap text
            wrap_text = cell.alignment.wrap_text if cell.alignment else False
            
            column = ColumnInfo(
                id=col_id,
                header=value,
                col_index=col,
                width=width,
                format=format_str,
                alignment=alignment,
                rowspan=rowspan,
                colspan=colspan,
                wrap_text=wrap_text
            )
            
            # Check for child columns (multi-row headers)
            if rowspan == 1 and colspan > 1:
                children = self._find_child_columns(worksheet, header_row + 1, col, colspan)
                column.children = children
            
            columns.append(column)
            processed_cols.add(col)
        
        return columns
    
    def _find_child_columns(self, worksheet: Worksheet, row: int, 
                            start_col: int, span: int) -> List[ColumnInfo]:
        """Find child columns under a parent header."""
        children = []
        for col in range(start_col, start_col + span):
            cell = worksheet.cell(row=row, column=col)
            value = self._get_cell_value(cell)
            if value:
                col_id = self._determine_column_id(value, col)
                format_str = self._determine_format(col_id, value)
                col_letter = get_column_letter(col)
                width = worksheet.column_dimensions[col_letter].width or 10
                
                children.append(ColumnInfo(
                    id=col_id,
                    header=value,
                    col_index=col,
                    width=width,
                    format=format_str
                ))
        return children
    
    def _determine_column_id(self, header_text: str, col_index: int) -> str:
        """Determine column ID from header text."""
        header_lower = header_text.lower().strip()
        
        # Check against known mappings
        for keyword, col_id in self.HEADER_MAPPINGS.items():
            if keyword in header_lower:
                return col_id
        
        # Fallback: generate ID from text
        clean_text = ''.join(c if c.isalnum() else '_' for c in header_lower)
        return f"col_{clean_text[:20]}"
    
    def _determine_format(self, col_id: str, header_text: str) -> str:
        """Determine number format for column."""
        # Numeric columns
        numeric_ids = {"col_qty_sf", "col_qty_pcs", "col_amount", "col_unit_price", 
                       "col_net", "col_gross", "col_cbm"}
        
        if col_id in numeric_ids:
            if "pcs" in col_id:
                return "#,##0"
            elif "cbm" in col_id:
                return "0.00"
            else:
                return "#,##0.00"
        
        return "@"  # Text format
    
    def _determine_data_source(self, sheet_name: str, columns: List[ColumnInfo]) -> str:
        """Determine what data source this sheet uses."""
        name_lower = sheet_name.lower()
        
        if any(s in name_lower for s in self.AGGREGATION_SHEETS):
            return "aggregation"
        elif any(s in name_lower for s in self.PROCESSED_SHEETS):
            return "processed_tables_multi"
        
        # Heuristic: if sheet has pcs/net/gross columns, it's likely packing list
        col_ids = {c.id for c in columns}
        if "col_qty_pcs" in col_ids or "col_net" in col_ids:
            return "processed_tables_multi"
        
        return "aggregation"
    
    def _extract_font_info(self, worksheet: Worksheet, row: int, col: int) -> Dict[str, Any]:
        """Extract font information from a cell."""
        cell = worksheet.cell(row=row, column=col)
        font = cell.font
        
        return {
            "name": font.name or "Times New Roman",
            "size": font.size or 12,
            "bold": font.bold or False,
            "italic": font.italic or False
        }
    
    def _extract_row_heights(self, worksheet: Worksheet, header_row: int) -> Dict[str, float]:
        """Extract row heights for header, data, and footer rows."""
        header_height = worksheet.row_dimensions[header_row].height or 30
        data_height = worksheet.row_dimensions[header_row + 1].height or 27
        
        return {
            "header": header_height,
            "data": data_height,
            "footer": header_height  # Usually same as header
        }
    
    def _check_multi_row_header(self, worksheet: Worksheet, header_row: int) -> bool:
        """Check if there's a multi-row header structure."""
        for merged in worksheet.merged_cells.ranges:
            if merged.min_row == header_row and merged.max_row > header_row:
                return True
        return False
    
    def _detect_static_content(self, worksheet: Worksheet, header_row: int, 
                               columns: List[ColumnInfo]) -> Dict[str, List[str]]:
        """Detect static content patterns in the data area."""
        hints = {}
        
        # Look for "Mark & Nº" type columns with static content
        for col in columns:
            if col.id == "col_static":
                static_values = []
                # Sample first few data rows
                for row in range(header_row + 1, min(header_row + 5, worksheet.max_row + 1)):
                    cell = worksheet.cell(row=row, column=col.col_index)
                    value = self._get_cell_value(cell)
                    if value and value not in static_values:
                        static_values.append(value)
                
                if static_values:
                    hints[col.id] = static_values
        
        return hints


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO)
    
    if len(sys.argv) < 2:
        print("Usage: python template_analyzer.py <template.xlsx>")
        sys.exit(1)
    
    analyzer = TemplateAnalyzer()
    result = analyzer.analyze_template(sys.argv[1])
    
    print(f"\nTemplate: {result.customer_code}")
    print(f"Sheets: {len(result.sheets)}")
    for sheet in result.sheets:
        print(f"\n  {sheet.name}:")
        print(f"    Header row: {sheet.header_row}")
        print(f"    Data source: {sheet.data_source}")
        print(f"    Columns: {len(sheet.columns)}")
        for col in sheet.columns:
            print(f"      - {col.id}: '{col.header}' (width={col.width:.1f})")
